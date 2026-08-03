"""Real data, webhook, and browser execution tools for Experiment 4-2."""

from __future__ import annotations

import hashlib
import json
import os
import re
import signal
import shutil
import subprocess
import tempfile
import time
from pathlib import Path
from typing import Any

import httpx
from config import Config


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _safe_output(path: str) -> Path:
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = Path(Config.WORKSPACE_DIR) / candidate
    candidate = candidate.resolve()
    candidate.relative_to(Path(Config.WORKSPACE_DIR).resolve())
    candidate.parent.mkdir(parents=True, exist_ok=True)
    return candidate


class ExtendedTools:
    async def excel_create_with_formula_and_screenshot(
        self, output_path: str, rows: list[dict[str, Any]]
    ) -> dict[str, Any]:
        """Create a real XLSX, apply formulas, and render a screenshot via LibreOffice."""
        from openpyxl import Workbook

        target = _safe_output(output_path)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invoice"
        sheet.append(["Item", "Quantity", "Unit price", "Total"])
        for index, row in enumerate(rows, 2):
            sheet.append([row["item"], float(row["quantity"]), float(row["unit_price"]),
                          f"=B{index}*C{index}"])
        total_row = len(rows) + 2
        sheet.cell(total_row, 3, "Grand total")
        sheet.cell(total_row, 4, f"=SUM(D2:D{total_row - 1})")
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 28
        for column in ("B", "C", "D"):
            sheet.column_dimensions[column].width = 16
        workbook.save(target)

        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            return {"success": False, "error": "LibreOffice is required for formula rendering"}
        started = time.perf_counter()
        process = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir",
             str(target.parent), str(target)],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120,
        )
        pdf = target.with_suffix(".pdf")
        if process.returncode != 0 or not pdf.is_file():
            return {"success": False, "error": process.stderr or process.stdout,
                    "returncode": process.returncode}
        import fitz

        document = fitz.open(pdf)
        screenshot = target.with_suffix(".png")
        document[0].get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False).save(screenshot)
        document.close()
        return {
            "success": True,
            "xlsx": {"path": str(target), "bytes": target.stat().st_size, "sha256": _sha(target)},
            "pdf": {"path": str(pdf), "bytes": pdf.stat().st_size, "sha256": _sha(pdf)},
            "screenshot": {"path": str(screenshot), "bytes": screenshot.stat().st_size,
                           "sha256": _sha(screenshot)},
            "formula_cells": [f"D{index}" for index in range(2, total_row + 1)],
            "rows": len(rows),
            "renderer": "LibreOffice headless + PyMuPDF",
            "latency_seconds": round(time.perf_counter() - started, 3),
        }

    async def webhook_post(self, url: str, payload: dict[str, Any]) -> dict[str, Any]:
        """POST JSON to a real HTTPS webhook and retain response evidence."""
        if not url.startswith("https://"):
            return {"success": False, "error": "Only HTTPS webhook URLs are allowed"}
        started = time.perf_counter()
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.post(url, json=payload)
        try:
            body = response.json()
        except ValueError:
            body = {"text": response.text[:2000]}
        return {
            "success": response.is_success,
            "status": response.status_code,
            "url": str(response.url),
            "response": body,
            "response_sha256": hashlib.sha256(response.content).hexdigest(),
            "response_bytes": len(response.content),
            "latency_seconds": round(time.perf_counter() - started, 3),
        }

    async def browser_navigate(self, url: str, screenshot_path: str) -> dict[str, Any]:
        """Navigate with real headless Chromium, extract content, and retain pixels."""
        if not url.startswith("https://"):
            return {"success": False, "error": "Only HTTPS URLs are allowed"}
        target = _safe_output(screenshot_path)
        started = time.perf_counter()
        from playwright.async_api import async_playwright

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=True)
            page = await browser.new_page(viewport={"width": 1280, "height": 720})
            response = await page.goto(url, wait_until="networkidle", timeout=60000)
            title = await page.title()
            text = (await page.locator("body").inner_text())[:4000]
            await page.screenshot(path=str(target), full_page=True)
            await browser.close()
        return {
            "success": bool(response and response.ok and target.is_file()),
            "url": url,
            "status": response.status if response else None,
            "title": title,
            "body_text": text,
            "screenshot": {"path": str(target), "bytes": target.stat().st_size,
                           "sha256": _sha(target)},
            "browser": "Chromium via Playwright",
            "latency_seconds": round(time.perf_counter() - started, 3),
        }

    async def virtual_desktop_execute(
        self, url: str, screenshot_path: str, expected_title: str | None = None
    ) -> dict[str, Any]:
        """Drive headful Chromium through X11 keyboard events and retain pixels."""
        if not url.startswith("https://"):
            return {"success": False, "error": "Only HTTPS URLs are allowed"}
        target = _safe_output(screenshot_path)
        required = {
            name: shutil.which(name)
            for name in ("Xvfb", "xdotool", "ffmpeg")
        }
        chromium = shutil.which("chromium") or shutil.which("chromium-browser")
        missing = [name for name, path in required.items() if not path]
        if not chromium:
            missing.append("chromium")
        if missing:
            return {"success": False, "error": f"Missing desktop executables: {missing}"}

        display_number = next((
            number for number in range(90, 130)
            if not Path(f"/tmp/.X11-unix/X{number}").exists()
            and not Path(f"/tmp/.X{number}-lock").exists()
        ), None)
        if display_number is None:
            return {"success": False, "error": "No free bounded X11 display number"}
        display = f":{display_number}"
        started = time.perf_counter()
        xvfb_process: subprocess.Popen[bytes] | None = None
        chromium_process: subprocess.Popen[bytes] | None = None

        def stop(process: subprocess.Popen[bytes] | None) -> None:
            if process is None or process.poll() is not None:
                return
            try:
                os.killpg(process.pid, signal.SIGTERM)
                process.wait(timeout=5)
            except (ProcessLookupError, subprocess.TimeoutExpired):
                try:
                    os.killpg(process.pid, signal.SIGKILL)
                except ProcessLookupError:
                    pass

        try:
            xvfb_process = subprocess.Popen(
                [required["Xvfb"], display, "-screen", "0", "1280x720x24", "-nolisten", "tcp"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                start_new_session=True,
            )
            socket_path = Path(f"/tmp/.X11-unix/X{display_number}")
            for _ in range(50):
                if socket_path.exists():
                    break
                if xvfb_process.poll() is not None:
                    return {"success": False, "error": "Xvfb exited before creating its socket"}
                time.sleep(0.1)
            else:
                return {"success": False, "error": "Xvfb did not become ready"}

            env = {**os.environ, "DISPLAY": display}
            with tempfile.TemporaryDirectory(prefix="exp4-computer-use-") as profile:
                chromium_process = subprocess.Popen(
                    [chromium, "--no-sandbox", "--disable-gpu", "--disable-dev-shm-usage",
                     f"--user-data-dir={profile}", "about:blank"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                    env=env, start_new_session=True,
                )
                window_id = ""
                for _ in range(100):
                    search = subprocess.run(
                        [required["xdotool"], "search", "--onlyvisible", "--class", "chromium"],
                        stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True, env=env,
                    )
                    if search.stdout.strip():
                        window_id = search.stdout.splitlines()[0].strip()
                        break
                    if chromium_process.poll() is not None:
                        return {"success": False, "error": "Chromium exited before opening a window"}
                    time.sleep(0.1)
                if not window_id:
                    return {"success": False, "error": "No visible Chromium window appeared"}

                subprocess.run(
                    [required["xdotool"], "windowfocus", "--sync", window_id],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, env=env,
                )
                input_receipts = []
                for command in (
                    [required["xdotool"], "key", "--window", window_id, "ctrl+l"],
                    [required["xdotool"], "type", "--window", window_id, "--delay", "15", url],
                    [required["xdotool"], "key", "--window", window_id, "Return"],
                ):
                    completed = subprocess.run(
                        command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, env=env,
                    )
                    input_receipts.append({"operation": command[1], "returncode": completed.returncode})
                    if completed.returncode != 0:
                        return {"success": False, "error": completed.stderr.strip(),
                                "input_receipts": input_receipts}

                title = ""
                for _ in range(100):
                    title_result = subprocess.run(
                        [required["xdotool"], "getwindowname", window_id],
                        stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True, env=env,
                    )
                    title = title_result.stdout.strip()
                    if title and (not expected_title or expected_title in title):
                        break
                    time.sleep(0.1)
                title_matched = bool(title and (not expected_title or expected_title in title))
                capture = subprocess.run(
                    [required["ffmpeg"], "-nostdin", "-loglevel", "error", "-f", "x11grab",
                     "-video_size", "1280x720", "-i", display, "-frames:v", "1", "-y", str(target)],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, env=env, timeout=30,
                )
                png_valid = target.is_file() and target.read_bytes().startswith(b"\x89PNG\r\n\x1a\n")
                return {
                    "success": capture.returncode == 0 and png_valid and title_matched,
                    "backend": "Xvfb + headful Chromium + xdotool",
                    "versions": {
                        "chromium": subprocess.run(
                            [chromium, "--version"], stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT, text=True,
                        ).stdout.strip(),
                        "xdotool": subprocess.run(
                            [required["xdotool"], "-v"], stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT, text=True,
                        ).stdout.strip(),
                        "ffmpeg": subprocess.run(
                            [required["ffmpeg"], "-version"], stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT, text=True,
                        ).stdout.splitlines()[0],
                    },
                    "display": display,
                    "window_id": window_id,
                    "url_entered_via_os_keyboard": url,
                    "window_title": title,
                    "expected_title": expected_title,
                    "expected_title_matched": title_matched,
                    "input_receipts": input_receipts,
                    "screenshot": ({"path": str(target), "bytes": target.stat().st_size,
                                    "sha256": _sha(target)} if png_valid else None),
                    "capture_returncode": capture.returncode,
                    "capture_error": capture.stderr.strip() or None,
                    "latency_seconds": round(time.perf_counter() - started, 3),
                }
        finally:
            stop(chromium_process)
            stop(xvfb_process)

    async def virtual_mobile_execute(
        self, container_name: str, screenshot_path: str
    ) -> dict[str, Any]:
        """Operate a real AndroidWorld emulator through ADB inside its container."""
        if not re.fullmatch(r"[A-Za-z0-9_.-]{1,128}", container_name):
            return {"success": False, "error": "Invalid Docker container name"}
        if not shutil.which("docker"):
            return {"success": False, "error": "Docker is required"}
        target = _safe_output(screenshot_path)
        started = time.perf_counter()

        def adb(*arguments: str, binary: bool = False) -> subprocess.CompletedProcess[Any]:
            return subprocess.run(
                ["docker", "exec", container_name, "adb", *arguments],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE,
                text=not binary, timeout=30,
            )

        inspect = subprocess.run(
            ["docker", "inspect", "--format", "{{.State.Running}} {{.Image}}", container_name],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
        )
        inspect_parts = inspect.stdout.strip().split(maxsplit=1)
        if inspect.returncode != 0 or not inspect_parts or inspect_parts[0] != "true":
            return {"success": False, "error": "AndroidWorld container is not running",
                    "container": container_name}
        boot = adb("shell", "getprop", "sys.boot_completed")
        devices = adb("devices", "-l")
        if boot.stdout.strip() != "1" or " device " not in f" {devices.stdout} ":
            return {"success": False, "error": "Android emulator is not boot-complete",
                    "container": container_name, "devices": devices.stdout.strip()}

        model = adb("shell", "getprop", "ro.product.model").stdout.strip()
        sdk = adb("shell", "getprop", "ro.build.version.sdk").stdout.strip()
        focus_before = adb("shell", "dumpsys", "window").stdout
        launch = adb("shell", "am", "start", "-W", "-a", "android.settings.WIFI_SETTINGS")
        focus_settings = adb("shell", "dumpsys", "window").stdout
        screenshot = adb("exec-out", "screencap", "-p", binary=True)
        if screenshot.returncode == 0:
            target.write_bytes(screenshot.stdout)
        home = adb("shell", "input", "keyevent", "KEYCODE_HOME")
        focus_home = adb("shell", "dumpsys", "window").stdout
        png_valid = target.is_file() and target.read_bytes().startswith(b"\x89PNG\r\n\x1a\n")
        settings_focused = "com.android.settings" in focus_settings
        launcher_focused = "launcher" in focus_home.lower()
        return {
            "success": all((launch.returncode == 0, settings_focused, home.returncode == 0,
                            launcher_focused, png_valid)),
            "backend": "AndroidWorld Docker emulator + ADB",
            "container": container_name,
            "container_image_id": inspect_parts[1] if len(inspect_parts) > 1 else None,
            "devices": devices.stdout.strip().splitlines(),
            "boot_completed": boot.stdout.strip(),
            "model": model,
            "api_level": sdk,
            "focus_before": next((line.strip() for line in focus_before.splitlines()
                                  if "mCurrentFocus=" in line), None),
            "settings_launch_returncode": launch.returncode,
            "settings_activity": next((line.strip() for line in launch.stdout.splitlines()
                                       if line.strip().startswith("Activity:")), None),
            "settings_focus": next((line.strip() for line in focus_settings.splitlines()
                                    if "mCurrentFocus=" in line), None),
            "home_input_returncode": home.returncode,
            "home_focus": next((line.strip() for line in focus_home.splitlines()
                                if "mCurrentFocus=" in line), None),
            "screenshot": ({"path": str(target), "bytes": target.stat().st_size,
                            "sha256": _sha(target)} if png_valid else None),
            "latency_seconds": round(time.perf_counter() - started, 3),
        }

    async def environment_capabilities(self) -> dict[str, Any]:
        """Report, without simulation, whether desktop/mobile backends are actually usable."""
        docker_image = subprocess.run(
            ["docker", "image", "inspect",
             "ghcr.io/anthropics/anthropic-quickstarts:computer-use-demo-latest"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        ).returncode == 0 if shutil.which("docker") else False
        android_container = os.getenv("ANDROID_WORLD_CONTAINER", "")
        active_devices: list[str] = []
        if android_container and re.fullmatch(r"[A-Za-z0-9_.-]{1,128}", android_container):
            devices = subprocess.run(
                ["docker", "exec", android_container, "adb", "devices", "-l"],
                stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True,
            ) if shutil.which("docker") else None
            if devices and devices.returncode == 0:
                active_devices = [line for line in devices.stdout.splitlines()[1:]
                                  if " device " in f" {line} "]
        return {
            "success": True,
            "computer_use_container_image_present": docker_image,
            "computer_use_host_stack_present": all(shutil.which(name) for name in
                                                    ("Xvfb", "xdotool", "ffmpeg"))
                                                    and bool(shutil.which("chromium") or
                                                             shutil.which("chromium-browser")),
            "computer_use_active_session": False,
            "android_world_container": android_container or None,
            "android_world_adb_present": bool(active_devices),
            "android_active_devices": active_devices,
            "note": "Availability probe only; execution gates are established by the dedicated desktop and mobile action receipts.",
        }
